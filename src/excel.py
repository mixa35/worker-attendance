"""Renders attendance.xlsx — one sheet per month, team-colored rows."""

from __future__ import annotations

import calendar
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import settings
from .db import connect

XLSX_PATH = Path(settings.data_dir) / "attendance.xlsx"
PRESENT_MARK = "✓"

COL_TEAM = 1
COL_ID = 2
COL_WORKER = 3
FIRST_DAY_COL = COL_WORKER + 1


def _sheet_name(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _open_or_create_workbook() -> Workbook:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        return load_workbook(XLSX_PATH)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def render_month_sheet(year: int, month: int) -> None:
    """Re-render the (year, month) sheet from current SQLite state. Idempotent."""
    days_in_month = calendar.monthrange(year, month)[1]
    month_prefix = f"{year:04d}-{month:02d}-"
    total_col = FIRST_DAY_COL + days_in_month

    wb = _open_or_create_workbook()
    name = _sheet_name(year, month)
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.cell(row=1, column=COL_TEAM, value="Team").font = bold
    ws.cell(row=1, column=COL_ID, value="National ID").font = bold
    ws.cell(row=1, column=COL_WORKER, value="Worker").font = bold
    for d in range(1, days_in_month + 1):
        c = ws.cell(row=1, column=FIRST_DAY_COL - 1 + d, value=d)
        c.font = bold
        c.alignment = center
    ws.cell(row=1, column=total_col, value="Total").font = bold

    conn = connect()
    try:
        teams = conn.execute(
            "SELECT id, name, color_hex FROM teams ORDER BY id"
        ).fetchall()
        row = 2
        for team in teams:
            fill = PatternFill(
                start_color=team["color_hex"],
                end_color=team["color_hex"],
                fill_type="solid",
            )
            workers = conn.execute(
                "SELECT id, name, "
                "       COALESCE(national_id, '#' || id) AS display_id "
                "FROM workers "
                "WHERE team_id = ? "
                "  AND (active = 1 OR id IN ("
                "      SELECT worker_id FROM attendance WHERE date LIKE ?"
                "  )) "
                "ORDER BY name",
                (team["id"], month_prefix + "%"),
            ).fetchall()
            for worker in workers:
                ws.cell(row=row, column=COL_TEAM, value=team["name"]).fill = fill
                id_cell = ws.cell(
                    row=row, column=COL_ID, value=worker["display_id"]
                )
                id_cell.fill = fill
                id_cell.alignment = center
                ws.cell(row=row, column=COL_WORKER, value=worker["name"]).fill = fill

                presence_rows = conn.execute(
                    "SELECT date, present FROM attendance "
                    "WHERE worker_id = ? AND date LIKE ?",
                    (worker["id"], month_prefix + "%"),
                ).fetchall()
                presence = {r["date"]: r["present"] for r in presence_rows}

                total = 0
                for d in range(1, days_in_month + 1):
                    date_iso = f"{month_prefix}{d:02d}"
                    cell = ws.cell(row=row, column=FIRST_DAY_COL - 1 + d)
                    cell.fill = fill
                    cell.alignment = center
                    if presence.get(date_iso) == 1:
                        cell.value = PRESENT_MARK
                        total += 1
                tcell = ws.cell(row=row, column=total_col, value=total)
                tcell.fill = fill
                tcell.alignment = center
                row += 1
    finally:
        conn.close()

    ws.column_dimensions[get_column_letter(COL_TEAM)].width = 16
    ws.column_dimensions[get_column_letter(COL_ID)].width = 14
    ws.column_dimensions[get_column_letter(COL_WORKER)].width = 22
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(FIRST_DAY_COL - 1 + d)].width = 4
    ws.column_dimensions[get_column_letter(total_col)].width = 8
    ws.freeze_panes = ws.cell(row=2, column=FIRST_DAY_COL).coordinate

    # Atomic write: save to a temp file in the same dir, then rename.
    # Prevents readers (e.g. /report, /download) from grabbing a half-written zip.
    import os
    tmp_path = XLSX_PATH.with_suffix(".xlsx.tmp")
    wb.save(tmp_path)
    os.replace(tmp_path, XLSX_PATH)
